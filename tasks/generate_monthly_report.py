"""
2026-04 淘顺实时授信月报生成脚本
"""
import openpyxl, os, matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
from docx import Document
from docx.shared import Inches, Pt, Cm, RGBColor
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.enum.table import WD_TABLE_ALIGNMENT
from docx.oxml.ns import qn
import datetime

# ============ Config ============
MONTH = '2026-04'
MONTH_CN = '4月'
DATA_DIR = r'F:\OneDrive - 湖南工商大学\codeworksplce\Work_code\Code_For_ts_risk\风控模型\数据文件\淘顺月报'
DATA_FILE = os.path.join(DATA_DIR, MONTH, '淘顺实时授信' + MONTH + '数据_bak2.xlsx')
TEMPLATE = os.path.join(DATA_DIR, '淘顺实时授信月报模板.docx')
OUTPUT_DIR = os.path.join(DATA_DIR, MONTH)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'淘顺实时授信月报-{MONTH_CN}月报.docx')
CHART_DIR = os.path.join(OUTPUT_DIR, 'charts')
os.makedirs(CHART_DIR, exist_ok=True)

# ============ Read Data ============
print('读取数据...')
wb = openpyxl.load_workbook(DATA_FILE, data_only=True)

# 整体统计表
rows_zt = [list(r) for r in wb['整体统计表'].iter_rows(values_only=True)]
zt_header = rows_zt[0]
zt_data = rows_zt[1:]

# vintage全量
rows_v = [list(r) for r in wb['vintage'].iter_rows(values_only=True)]
v_header = rows_v[0]
v_data = [r for r in rows_v[1:] if len(r) > 1 and r[1] == '全国']

# vintage全量预测
rows_vp = [list(r) for r in wb['vintage-pre'].iter_rows(values_only=True)]
vp_data = [r for r in rows_vp[1:] if len(r) > 1 and r[1] == '全国']

# vintage公众
rows_vg = [list(r) for r in wb['vintage-gz'].iter_rows(values_only=True)]
vg_data = [r for r in rows_vg[1:] if len(r) > 1 and r[1] == '全国']

# vintage公众预测
rows_vgp = [list(r) for r in wb['vintage-gz-pre'].iter_rows(values_only=True)]
vgp_data = [r for r in rows_vgp[1:] if len(r) > 1 and r[1] == '全国']

wb.close()

print(f'整体统计表: {len(zt_data)}行')
print(f'vintage全量: {len(v_data)}行, 预测: {len(vp_data)}行')
print(f'vintage公众: {len(vg_data)}行, 预测: {len(vgp_data)}行')

# ============ Helper Functions ============
def get_mob_cols(header):
    """获取mob列索引"""
    mobs = {}
    for i, h in enumerate(header):
        if isinstance(h, str) and h.lower().startswith('mob'):
            mobs[h] = i
    return mobs

def get_final_pred(row, mob_cols):
    """取方法二的终损预测值（最后一个非空mob值）"""
    vals = []
    for m, idx in sorted(mob_cols.items(), key=lambda x: int(x[0][3:])):
        if idx < len(row) and row[idx] is not None:
            v = float(row[idx]) if row[idx] not in ('', None) else None
            if v is not None:
                vals.append(v)
    return vals[-1] * 100 if vals else None  # 转百分比

# ============ Generate Vintage Charts ============
print('\n生成vintage图表...')

def plot_vintage(data, pred_data, title, filename, mob_limit=15):
    """绘制vintage曲线图"""
    mob_cols = get_mob_cols(v_header)
    
    fig, ax = plt.subplots(figsize=(12, 6))
    
    # 只画最近12期
    recent = data[-12:] if len(data) > 12 else data
    
    colors = plt.cm.Blues(np.linspace(0.3, 0.9, len(recent)))
    for i, row in enumerate(recent):
        order_month = row[0]
        mob_vals = []
        mob_nums = []
        for m, idx in sorted(mob_cols.items(), key=lambda x: int(x[0][3:])):
            mob_n = int(m[3:])
            if mob_n > mob_limit: break
            if idx < len(row) and row[idx] is not None:
                v = float(row[idx]) if row[idx] not in ('', None) else None
                if v is not None:
                    mob_vals.append(v * 100)
                    mob_nums.append(mob_n)
        color = '#2196F3' if '2026' in str(order_month) else '#90CAF9'
        alpha = 0.9 if '2026' in str(order_month) else 0.3
        ax.plot(mob_nums, mob_vals, 'o-', color=color, alpha=alpha, linewidth=1.5,
                label=str(order_month))
    
    # 预测终损值标注
    if pred_data:
        last_pred = pred_data[-1] if pred_data else None
        if last_pred:
            pred_val = get_final_pred(last_pred, mob_cols)
            order_month = last_pred[0]
            if pred_val:
                ax.axhline(y=pred_val, color='red', linestyle='--', alpha=0.5, linewidth=1)
                ax.text(mob_limit - 2, pred_val, f'终损预估: {pred_val:.1f}%', 
                        color='red', fontsize=10, fontweight='bold',
                        bbox=dict(facecolor='white', alpha=0.7))
    
    ax.set_xlabel('账期(MOB)', fontsize=12)
    ax.set_ylabel('逾期率(%)', fontsize=12)
    ax.set_title(title, fontsize=14, fontweight='bold')
    ax.legend(loc='upper left', fontsize=8, ncol=2)
    ax.grid(True, alpha=0.3)
    ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.1f%%'))
    
    plt.tight_layout()
    fig.savefig(filename, dpi=200, bbox_inches='tight')
    plt.close(fig)
    return filename

# 全量vintage
plot_vintage(v_data, vp_data, '淘顺实时授信 - 全量用户Vintage曲线', 
             os.path.join(CHART_DIR, 'vintage_all.png'))
# 公众vintage
plot_vintage(vg_data, vgp_data, '淘顺实时授信 - 公众用户Vintage曲线',
             os.path.join(CHART_DIR, 'vintage_gz.png'))

print('vintage图表生成完成')

# ============ Generate Report ============
print('\n生成本月数据概览...')

# 计算全国和各省概览
overview_text = '淘顺实时授信截至2026年4月底，'
total_orders = 0
total_overdue = 0
for r in zt_data:
    if r[0] == '全国':
        total_orders = int(r[5]) if r[5] else 0
        total_overdue = int(r[4]) if r[4] else 0
        overview_text += f'累计办单{total_orders/10000:.1f}万笔，逾期率为{float(r[6].strip("%")):.2f}%，'
        overview_text += f'金额逾期率为{float(r[9].strip("%")):.2f}%。'
        break

print(overview_text)

# 各省逾期率排序
prov_overdue = []
for r in zt_data:
    if r[0] != '全国':
        try:
            rate = float(r[6].strip('%'))
            prov_overdue.append((r[0], rate, int(r[5])))
        except:
            pass
prov_overdue.sort(key=lambda x: x[1], reverse=True)

print('各省逾期率(Top5):')
for p, r, n in prov_overdue[:5]:
    print(f'  {p}: {r:.2f}% ({n:,}单)')

# 新增：与3月对比分析（读3月数据）
mar_path = os.path.join(DATA_DIR, '2026-03', '淘顺实时授信2026-03数据.xlsx')
mar_overdue_rate = None
if os.path.exists(mar_path):
    wb3 = openpyxl.load_workbook(mar_path, data_only=True)
    ws3 = wb3['整体统计表']
    for r in ws3.iter_rows(values_only=True):
        if r[0] == '全国':
            mar_overdue_rate = float(r[6].strip('%'))
            break
    wb3.close()

if mar_overdue_rate:
    cur_rate = float(zt_data[0][6].strip('%')) if zt_data[0][0] == '全国' else None
    if cur_rate:
        diff = cur_rate - mar_overdue_rate
        print(f'逾期率变化: {mar_overdue_rate:.2f}% → {cur_rate:.2f}% ({diff:+.2f}pp)')

# ============ Vintage终损预估对比 ============
print('\n终损预估对比（预测Method 2）：')
def get_pred_summary(pred_data):
    mob_cols = get_mob_cols(v_header)
    results = {}
    for row in pred_data:
        order_month = str(row[0])
        if order_month >= '2025-01':
            pred_val = get_final_pred(row, mob_cols)
            # 取最新实际mob
            latest_actual = None
            for m, idx in sorted(mob_cols.items(), key=lambda x: int(x[0][3:])):
                if idx < len(row) and row[idx] is not None and str(row[idx]) not in ('', 'None'):
                    latest_actual = float(row[idx]) * 100
            results[order_month] = {'pred': pred_val, 'latest': latest_actual}
    return results

all_pred = get_pred_summary(vp_data)
gz_pred = get_pred_summary(vgp_data)

for m, vals in sorted(all_pred.items())[-6:]:
    print(f'  {m}: 终损={vals["pred"]:.1f}%, 最新实际={vals["latest"]:.1f}%')

# ============ 重点地市分析 ============
print('\n重点地市逾期率Top:')
# 从地市逾期情况分析
ws_city = openpyxl.load_workbook(DATA_FILE, data_only=True)['地市逾期情况']
city_rows = [list(r) for r in ws_city.iter_rows(values_only=True)]
high_risk_cities = []
for r in city_rows[1:]:
    if r[0] not in ('全国', '省份') and r[4]:
        try:
            rate = float(r[4].strip('%'))
            orders = int(r[3])
            if orders >= 100 and rate >= 8:
                high_risk_cities.append((r[0], r[1], rate, orders))
        except:
            pass
high_risk_cities.sort(key=lambda x: x[2], reverse=True)
for p, c, r, n in high_risk_cities[:10]:
    print(f'  {c}({p}): {r:.2f}% ({n}单)')

print('\n✅ 数据读取完成，准备构建月报')
